import tkinter as tk
import tkinter.ttk as ttk
import os
import threading
from tkinter.filedialog import asksaveasfilename
from woocommerce import API
from openpyxl import load_workbook
from datetime import datetime, date
from time import sleep
import sys
from os import path

# Constants
BULLET = '\u2022'
__version__ = '0.0.1'
TEHRAN = 'تهران'
PERSIAN_NUMBERS = u'۱۲۳۴۵۶۷۸۹۰'
ARABIC_NUMBERS = u'١٢٣٤٥٦٧٨٩٠'
ENGLISH_NUMBERS = u'1234567890'
PERSIAN_TRANS = str.maketrans(PERSIAN_NUMBERS, ENGLISH_NUMBERS)
ARABIC_TRANS = str.maketrans(ARABIC_NUMBERS, ENGLISH_NUMBERS)
io_thread = None
progress_thread = None


def get_api_settings():
    consumer_key = os.environ.get('PISHRO_CONSUMER_KEY', '')
    consumer_secret = os.environ.get('PISHRO_CONSUMER_SECRET', '')
    ent_consumer_key.delete(0, tk.END)
    ent_consumer_secret.delete(0, tk.END)
    ent_consumer_key.insert(0, consumer_key)
    ent_consumer_secret.insert(0, consumer_secret)


def set_output_file_entry_text(text):
    ent_output_path.delete(0, tk.END)
    ent_output_path.insert(0, text)


def choose_save_output_file():
    filetypes = (
        ('Excel File', '*.xlsx'),
    )
    save_file_path = asksaveasfilename(
        defaultextension='.xlsx',
        title='لطفا محل ذخیره را انتخاب کنیدپ',
        filetypes=filetypes,
        initialdir=path.split(ent_output_path.get())[0],
        initialfile=path.split(ent_output_path.get())[1]
    )
    if save_file_path:
        set_output_file_entry_text(save_file_path)


def show_hide_api_settings():
    if ent_consumer_key['show'] == BULLET:
        ent_consumer_key.config(show='')
        ent_consumer_secret.config(show='')
        btn_show_api_settings['text'] = 'مخفی کردن مشخصات'
    else:
        ent_consumer_key.config(show=BULLET)
        ent_consumer_secret.config(show=BULLET)
        btn_show_api_settings['text'] = 'نمایش مشخصات'


def get_woocommerce_object():
    return API(
        ent_website_address.get(),
        consumer_key=ent_consumer_key.get(),
        consumer_secret=ent_consumer_secret.get(),
        user_agent=f'Pishro Generator{__version__}',
        timeout=150,
    )


def get_all_printing_orders():
    filters = {
        'per_page': 100,
        'page': 1,
        'status': 'processing'
    }
    wcapi = get_woocommerce_object()
    order_list = []
    result = wcapi.get('orders', params=filters).json()
    order_list += result
    while not len(result) < filters['per_page']:
        filters['page'] += 1
        result = wcapi.get('orders', params=filters).json()
        order_list += result
    return order_list


def get_none_empty(*args):
    for item in args:
        if item:
            return item
    return ''


def generate_pishro_excel():
    global io_thread
    pb_generate['value'] = 0
    label_generate_status['text'] = 'در حال دریافت سفارشات'
    orders = get_all_printing_orders()
    out_wb = load_workbook(path.abspath(
        path.join(path.dirname(__file__), 'assets/sample.xlsx')))
    ws = out_wb.active
    date = get_date()
    pb_generate['mode'] = 'determinate'
    label_generate_status['text'] = 'ایجاد فایل اکسل'
    for index, order in enumerate(orders):
        index += 2
        address_billing = order['billing']
        address_shipping = order['shipping']
        ws[f'A{index}'] = str(date)
        ws[f'B{index}'] = ent_pishro_customer_code.get()
        ws[f'C{index}'] = order['id']
        ws[f'D{index}'] = ''
        ws[f'E{index}'] = '1'
        ws[f'F{index}'] = '1'
        ws[f'G{index}'] = len(order['line_items'])
        ws[f'H{index}'] = str(int(order['total']) * 10)
        ws[f'I{index}'] = '0'
        ws[f'J{index}'] = get_none_empty(
            address_shipping['phone'], address_billing['phone']).translate(PERSIAN_TRANS).translate(ARABIC_TRANS)
        ws[f'K{index}'] = get_none_empty(
            address_shipping['postcode'], address_billing['postcode']).translate(PERSIAN_TRANS).translate(ARABIC_TRANS)
        ws[f'L{index}'] = ' '.join((get_none_empty(
            address_shipping['first_name'], address_billing['first_name']), get_none_empty(
            address_shipping['last_name'], address_billing['last_name']))).strip()
        ws[f'M{index}'] = get_none_empty(
            address_shipping['city'], address_billing['city'])
        ws[f'N{index}'] = ' '.join((get_none_empty(
            address_shipping['address_1'], address_billing['address_1']), get_none_empty(
            address_shipping['address_2'], address_billing['address_2']))).strip()
        ws[f'O{index}'] = 'SP'
        ws[f'P{index}'] = 'آرایشی بهداشتی'
        ws[f'R{index}'] = '6584'
        ws[f'U{index}'] = order['customer_note'].strip()
        pb_generate['value'] = (index - 1)/len(orders) * 100
    if not len(orders):
        pb_generate['value'] = 1000
        label_generate_status['text'] = 'سفارشی موجود نبود'
    else:
        out_wb.save(ent_output_path.get())
        label_generate_status['text'] = 'پایان'
    io_thread = None


def generate_pishro_worker_thread():
    global io_thread, progress_thread
    if not io_thread:
        io_thread = threading.Thread(target=generate_pishro_excel)
        io_thread.start()
        progress_thread = threading.Thread(target=load_progress)
        progress_thread.start()


def load_progress():
    global progress_thread
    pb_generate['mode'] = 'indeterminate'
    a = 'f' + str(pb_generate['mode'])
    while pb_generate['mode'] == 'indeterminate':
        pb_generate['value'] += 1
        sleep(0.001)

    progress_thread = None


def get_date():
    return date(year=int(ent_date_year.get()),
                month=int(ent_date_month.get()), day=int(ent_date_day.get()))


def year_validate(P):
    return ((len(P)) <= 4 and P.isdigit()) or len(P) == 0


def day_month_validate(P):
    return ((len(P)) <= 2 and P.isdigit()) or len(P) == 0


def set_today_date():
    date = datetime.now().date()
    ent_date_year.delete(0, tk.END)
    ent_date_month.delete(0, tk.END)
    ent_date_day.delete(0, tk.END)
    ent_date_year.insert(0, date.year)
    ent_date_month.insert(0, date.month)
    ent_date_day.insert(0,  date.day)


def set_ent_output_default_path():
    set_output_file_entry_text(
        path.join(path.dirname(sys.argv[0]), 'out.xlsx'))


window = tk.Tk()
window.title('ایجاد کننده اکسل پیشرو')
window.minsize(width=381, height=1)
window.resizable(width=True, height=False)

year_validator = (window.register(year_validate), '%P')
day_month_validator = (window.register(day_month_validate), '%P')

frm_parent = tk.Frame(relief=tk.GROOVE, borderwidth=2)
frm_parent.pack(padx=5, pady=5, fill=tk.X)


# Input Section
frm_api_settings = ttk.Frame(
    master=frm_parent, relief=tk.GROOVE, borderwidth=2)
frm_api_settings['padding'] = 5
frm_api_settings.pack(padx=10, pady=10, fill=tk.X)
frm_api_settings.columnconfigure(1, weight=1)


lbl_consumer_key = tk.Label(text='consumer_key', master=frm_api_settings)
ent_consumer_key = tk.Entry(master=frm_api_settings, width=35,)
ent_consumer_key.grid(pady=2, row=0, column=1, columnspan=2, sticky='ew', )
lbl_consumer_key.grid(pady=2, row=0, column=0, sticky='w')
ent_consumer_key.config(show=BULLET)

lbl_consumer_secert = tk.Label(
    text='consumer_secret', master=frm_api_settings,)
ent_consumer_secret = tk.Entry(master=frm_api_settings, width=35,)
ent_consumer_secret.grid(pady=2, row=1, column=1, columnspan=2, sticky='ew')
lbl_consumer_secert.grid(pady=2, padx=(0, 5), row=1, column=0, sticky='w')
ent_consumer_secret.config(show=BULLET)


lbl_website_address = tk.Label(
    text='Website address', master=frm_api_settings,)
ent_website_address = tk.Entry(master=frm_api_settings, width=35,)
ent_website_address.grid(pady=2, row=2, column=1, columnspan=2, sticky='ew')
lbl_website_address.grid(pady=2, padx=(0, 5), row=2, column=0, sticky='w')

lbl_pishro_customer_code = tk.Label(
    text='Pishro customer code', master=frm_api_settings,)
ent_pishro_customer_code = tk.Entry(master=frm_api_settings, width=35,)
ent_pishro_customer_code.grid(pady=2, row=3, column=1, columnspan=2, sticky='ew')
lbl_pishro_customer_code.grid(pady=2, padx=(0, 5), row=3, column=0, sticky='w')

frm_api_settings_actions = tk.Frame(master=frm_api_settings)
frm_api_settings_actions.grid(
    row=4, column=1, columnspan=2, pady=5, sticky='e')

btn_show_api_settings = tk.Button(
    master=frm_api_settings_actions, text='نمایش مشخصات', command=show_hide_api_settings)
btn_show_api_settings.grid(row=0, column=1, padx=(2.5, 0), sticky='e')
btn_get_api_settings = tk.Button(
    master=frm_api_settings_actions, text='خواندن مشخصات', command=get_api_settings)
btn_get_api_settings.grid(row=0, column=0, padx=2.5, sticky='e')
get_api_settings()

lbl_date = tk.Label(master=frm_api_settings, text='تاریخ تحویل')
frm_date = tk.Label(master=frm_api_settings)

lbl_date_year = tk.Label(master=frm_date, text='سال')
lbl_date_month = tk.Label(master=frm_date, text='ماه')
lbl_date_day = tk.Label(master=frm_date, text='روز')

ent_date_year = tk.Entry(master=frm_date, width=4,
                         validate='key', validatecommand=year_validator)
ent_date_month = tk.Entry(master=frm_date, width=2,
                          validate='key', validatecommand=day_month_validator)
ent_date_day = tk.Entry(master=frm_date, width=2,
                        validate='key', validatecommand=day_month_validator)

btn_date_today = tk.Button(master=frm_api_settings,
                           text='برو به امروز', command=set_today_date)

lbl_date.grid(row=5, column=2, sticky='e')
btn_date_today.grid(row=4, column=0, sticky='w')
frm_date.grid(row=5, column=1, sticky='e')

ent_date_year.grid(row=0, column=0, padx=1.25)
lbl_date_year.grid(row=0, column=1, padx=1.25)
ent_date_month.grid(row=0, column=2, padx=1.25)
lbl_date_month.grid(row=0, column=3, padx=1.25)
ent_date_day.grid(row=0, column=4,  padx=1.25)
lbl_date_day.grid(row=0, column=5,  padx=1.25)

set_today_date()

# OutPut section
frm_output_file = ttk.Frame(master=frm_parent, relief=tk.GROOVE, borderwidth=2)
frm_output_file['padding'] = 5
frm_output_file.pack(padx=10, pady=10, fill=tk.X)
frm_output_file.columnconfigure(1, weight=1)
btn_save_output_file = tk.Button(
    text='انتخاب محل ذخیره', command=choose_save_output_file, master=frm_output_file)
lbl_output = tk.Label(text='آدرس فایل خروجی', master=frm_output_file,)
sv_output_excel = tk.StringVar()
ent_output_path = tk.Entry(master=frm_output_file, width=35,
                           textvariable=sv_output_excel)
btn_save_output_file.grid(pady=2, row=0, column=0)
ent_output_path.grid(pady=2, row=0, column=1, padx=5, sticky='ew')
lbl_output.grid(pady=2, row=0, column=2, sticky='e')


set_ent_output_default_path()

pb_generate = ttk.Progressbar(
    master=frm_output_file, orient=tk.HORIZONTAL, length=100, mode='determinate')
pb_generate.grid(row=1, column=0, columnspan=2, sticky='ew')

label_generate_status = tk.Label(
    master=frm_output_file, text='آغاز نشده')
label_generate_status.grid(row=1, column=2, sticky='e')


frm_ouput_action = tk.Frame(master=frm_output_file)
frm_ouput_action.grid(row=2, column=1, columnspan=2, pady=5, sticky='e')

btn_generate_output = tk.Button(
    master=frm_ouput_action, text='ایجاد خروجی', command=generate_pishro_worker_thread)
btn_generate_output.grid(row=0, column=0, padx=(2.5, 0), sticky='e')
window.mainloop()
